import asyncio
import discord
import datetime
import os
import openpyxl
import requests
from bs4 import BeautifulSoup
import random

client=discord.Client()

access_token=os.environ["BOT_TOKEN"]


ran=random.randint(0,5)
role=["마피아","시민","의사","경찰","군인"]

yes = "⭕"
no = "❌"
ok="👍"
reactions = [yes, no, ok]




@client.event
async def on_ready():
    print(client.user.name)
    print('started bot')    
    await client.change_presence(activity=discord.Activity(type=discord.ActivityType.custom, name="정답 받는중"))

@client.event
async def on_member_join(member):
        print(member + "님 서버에 들어온걸 환영합니다.")

@client.event
async def on_message(message):
    money=0
    an=0
    b=0
    c=0
    d=0
    e=0
    def member():
        file = openpyxl.load_workbook("memberlist.xlsx")
        wb = file.active
   
        for i in range(1, 101):
            return wb["A" + str(i)].value == str(message.author.id)

    def quiz():
        file = openpyxl.load_workbook("memberlist.xlsx")
        wb = file.active
   
        for i in range(1, 101):
            return wb["D" + str(i)].value == str(message.author.id)


    def check(reaction, user):
        return user == message.author and str(reaction.emoji) in reactions

    def check2(reaction, user):
        return  str(reaction.emoji) in reactions
    
        

    if message.content.startswith("!"):
        if message.content == "!시간표":
            today=datetime.datetime.now().weekday()-9
            if today==0:
                emb=discord.Embed(color=0xff9900)
                emb.add_field(name="오늘의 시간표",value="월요일",inline=False)
                emb.add_field(name="조회",value="https://zoom.us/j/4620065666?pwd=aDJ1T2ZSQkRaSVFBaEczSmtMVk9CZz09 (462 006 5666, 1357)")
                emb.add_field(name="1교시", value="[국어](https://zoom.us/j/3687297166?pwd=UVA3bDJ0VklhUzZVNGlBRXVUdk1UQT09) 368 729 7166, 1004", inline=False)           
                emb.add_field(name="2교시", value="[영어](https://zoom.us/j/6924056614?pwd=WGR1SGsrU0xhVlFXcWFKbjlmcExxUT09) 692 405 6614, 1212", inline=False)
                emb.add_field(name="3교시", value="[음악](https://zoom.us/j/8992587200?pwd=WnRDY0ZLUUJPdTE0VWMrVTdCUHlQQT09) 899 258 7200, 1243",inline=False)
                emb.add_field(name="4교시", value="[도덕](https://us02web.zoom.us/j/6725551179?pwd=QjBsYWh4cGlVRUNtcnJyeGxQNEpYZz09) 672 555 1179, 12345",inline=False)
                emb.add_field(name="5교시", value="[역사](https://zoom.us/j/6404846050?pwd=MklheVFUOUlnclVFd1l5OXlxRnBTUT09) 640 484 6050, durian",inline=False)
                emb.add_field(name="6교시", value="[과학](https://zoom.us/j/3440254414?pwd=alNVQ0NlS0VYdkZoRWlLeWhpS2thQT09) 344 025 4414, 1234",inline=False)
                await message.channel.send(embed=emb)
            elif today==1:
                emb=discord.Embed(color=0xff9900)
                emb.add_field(name="오늘의 시간표",value="화요일",inline=False)
                emb.add_field(name="조회",value="https://zoom.us/j/4620065666?pwd=aDJ1T2ZSQkRaSVFBaEczSmtMVk9CZz09 (462 006 5666, 1357)")
                emb.add_field(name="1교시", value="[과학](https://zoom.us/j/3440254414?pwd=alNVQ0NlS0VYdkZoRWlLeWhpS2thQT09) 344 025 4414, 1234", inline=False)
                emb.add_field(name="2교시", value="[수학](https://us02web.zoom.us/j/6718992309?pwd=ci8rQVByUXVPMElwQ01ITlBwUmxLdz09) 671 899 2309, 486486", inline=False)
                emb.add_field(name="3교시", value="[국어](https://zoom.us/j/3687297166?pwd=UVA3bDJ0VklhUzZVNGlBRXVUdk1UQT09) 368 729 7166, 1004",inline=False)           
                emb.add_field(name="4교시", value="[역사](https://zoom.us/j/6404846050?pwd=MklheVFUOUlnclVFd1l5OXlxRnBTUT09) 640 484 6050, durian",inline=False)
                emb.add_field(name="5교시", value="[미술](https://zoom.us/j/2210779923?pwd=MHlPMjJkYnV2QmhZRVdRZWk1bnBxdz09) 221 077 9923, 2121",inline=False)
                emb.add_field(name="6교시", value="[한문](https://zoom.us/j/4620065666?pwd=aDJ1T2ZSQkRaSVFBaEczSmtMVk9CZz09)462 006 5666, 1357",inline=False)
                emb.add_field(name="7교시", value="[영어](https://zoom.us/j/6924056614?pwd=WGR1SGsrU0xhVlFXcWFKbjlmcExxUT09) 692 405 6614, 1212",inline=False)
                await message.channel.send(embed=emb)
            elif today==2:
                emb=discord.Embed(color=0xff9900)
                emb.add_field(name="오늘의 시간표",value="수요일",inline=False)
                emb.add_field(name="조회",value="https://zoom.us/j/4620065666?pwd=aDJ1T2ZSQkRaSVFBaEczSmtMVk9CZz09 (462 006 5666, 1357)")
                emb.add_field(name="1교시", value="[영어](https://zoom.us/j/6924056614?pwd=WGR1SGsrU0xhVlFXcWFKbjlmcExxUT09) 692 405 6614, 1212", inline=False)
                emb.add_field(name="2교시", value="[체육](https://www.ebsoc.co.kr/) 온라인 클래스", inline=False)
                emb.add_field(name="3교시", value="[미술](https://zoom.us/j/2210779923?pwd=MHlPMjJkYnV2QmhZRVdRZWk1bnBxdz09) 221 077 9923, 2121o",inline=False)
                emb.add_field(name="4교시", value="[수학](https://us02web.zoom.us/j/6718992309?pwd=ci8rQVByUXVPMElwQ01ITlBwUmxLdz09) 671 899 2309, 486486",inline=False)
                emb.add_field(name="5교시", value="[도덕](https://us02web.zoom.us/j/6725551179?pwd=QjBsYWh4cGlVRUNtcnJyeGxQNEpYZz09) 672 555 1179, 12345",inline=False)
                emb.add_field(name="6교시", value="[스포츠](https://sites.google.com/ggh.goe.go.kr/eovud2/%ED%99%88/%EC%97%AD%EC%82%AC) 각자 맞는 선생님 방에 접속",inline=False)
                await message.channel.send(embed=emb)
            elif today==3:
                emb=discord.Embed(color=0xff9900)
                emb.add_field(name="오늘의 시간표",value="목요일",inline=False)
                emb.add_field(name="조회",value="https://zoom.us/j/4620065666?pwd=aDJ1T2ZSQkRaSVFBaEczSmtMVk9CZz09 (462 006 5666, 1357)")
                emb.add_field(name="1교시", value="[국어](https://zoom.us/j/3687297166?pwd=UVA3bDJ0VklhUzZVNGlBRXVUdk1UQT09) 368 729 7166, 1004", inline=False)            
                emb.add_field(name="2교시", value="[영어](https://zoom.us/j/6924056614?pwd=WGR1SGsrU0xhVlFXcWFKbjlmcExxUT09) 692 405 6614, 1212", inline=False)
                emb.add_field(name="3교시", value="[수학](https://us02web.zoom.us/j/6718992309?pwd=ci8rQVByUXVPMElwQ01ITlBwUmxLdz09) 671 899 2309, 486486",inline=False)
                emb.add_field(name="4교시", value="[체육](https://www.ebsoc.co.kr/) 온라인 클래스",inline=False)
                emb.add_field(name="5교시", value="[자율](https://zoom.us/j/4620065666?pwd=aDJ1T2ZSQkRaSVFBaEczSmtMVk9CZz09) 462 006 5666, 1357",inline=False)
                emb.add_field(name="6교시", value="[과학](https://zoom.us/j/3440254414?pwd=alNVQ0NlS0VYdkZoRWlLeWhpS2thQT09) 344 025 4414, 1234",inline=False)
                emb.add_field(name="7교시", value="[정보](https://zoom.us/j/7525745591?pwd=YjlRQ1JWNDFxM1lsUUdmNEZraHhLZz09) 752 574 5591, com2",inline=False)
                await message.channel.send(embed=emb)
            elif today==4:
                emb=discord.Embed(color=0xff9900)
                emb.add_field(name="오늘의 시간표",value="금요일",inline=False)
                emb.add_field(name="조회",value="https://zoom.us/j/4620065666?pwd=aDJ1T2ZSQkRaSVFBaEczSmtMVk9CZz09 (462 006 5666, 1357)")
                emb.add_field(name="1교시", value="[체육](https://www.ebsoc.co.kr/) 온라인 클래스", inline=False)
                emb.add_field(name="2교시", value="[국어](https://zoom.us/j/3687297166?pwd=UVA3bDJ0VklhUzZVNGlBRXVUdk1UQT09) 368 729 7166, 1004", inline=False)            
                emb.add_field(name="3교시", value="[한문](https://zoom.us/j/4620065666?pwd=aDJ1T2ZSQkRaSVFBaEczSmtMVk9CZz09) 462 006 5666, 1357",inline=False)
                emb.add_field(name="4교시", value="[수학](https://us02web.zoom.us/j/6718992309?pwd=ci8rQVByUXVPMElwQ01ITlBwUmxLdz09) 671 899 2309, 486486",inline=False)
                emb.add_field(name="5교시", value="[과학](https://zoom.us/j/3440254414?pwd=alNVQ0NlS0VYdkZoRWlLeWhpS2thQT09) 344 025 4414, 1234",inline=False)
                emb.add_field(name="6교시", value="[역사](https://zoom.us/j/6404846050?pwd=MklheVFUOUlnclVFd1l5OXlxRnBTUT09) 640 484 6050, durian",inline=False)
                await message.channel.send(embed=emb)
            else:
                await message.channel.send("오늘은 등교일이 아닙니다")

        elif message.content=="!도움":
            em=discord.Embed(color=0xff9900)
            em.add_field(name="!시간표",value="오늘의 시간표를 보여주는 명령입니다.(수업 접속링크도 함께)",inline=False)
            em.add_field(name="!급식",value="오늘의 급식을 안내하는 명령어 입니다.",inline=False)
            em.add_field(name="!가입",value="회원가입을 진행하는 명령어입니다.",inline=False)
            em.add_field(name="!탈퇴",value="가입된 회원정보를 삭제하고 탈퇴하는 명령어 입니다.",inline=False)
            em.add_field(name="!내정보",value="이름, 서버닉네임, 가입일을 안내하는 명령어 입니다.",inline=False)
            em.add_field(name="!날씨", value="오늘의 날씨를 알려주는 명령어 입니다.",inline=False)
            em.add_field(name="!청소",value="메시지를 삭제하는 명령어 입니다.(ex> !청소 20)",inline=False)
            em.add_field(name="!돈",value="현재 보유중인 돈의 양을 알려주는 명령어 입니다.", inline=False)
            em.add_field(name="!돈받기",value="돈을 받는 명령어 입니다.",inline=False)

            await message.channel.send(embed=em)

        elif message.content=="!급식":
            url1="https://daepyong-m.goesw.kr/index.do"
            res1=requests.get(url1)
            res1.raise_for_status()
            soup1=BeautifulSoup(res1.text,"lxml")
            sou=soup1.find(attrs={"class":"txt_line"})

            await message.channel.send(sou.get_text)

        
    
            
        elif message.content == "!가입":
            file = openpyxl.load_workbook("memberlist.xlsx")
            wb = file.active
            for i in range(1, 101):
                if wb["A" + str(i)].value == str(message.author.id):
                    await message.channel.send("이미 가입된 사용자입니다.")
                    break

                elif wb["A" + str(i)].value == None:
                    embed = discord.Embed(title="회원 가입을 진행합니다", description="아래의 이모지에 반응하세요.",color=0xff9900)
                    msg = await message.channel.send(embed=embed)
                    await msg.add_reaction(yes)
                    await msg.add_reaction(no)

                    try:
                        reaction, user=await client.wait_for("reaction_add", check=check, timeout=15)
                    
                        if str(reaction.emoji) == yes:
                            await msg.delete()                

                            wb["A" + str(i)].value = str(message.author.id)
                            wb["B" + str(i)].value = str(message.author)
                            wb["C" + str(i)].value = str(datetime.datetime.now().replace(microsecond=0))
                            wb["D" + str(i)].value = money
                            await message.channel.send("가입이 완료되었습니다.")
                            break
            
                        elif str(reaction.emoji) == no:
                            await msg.delete()
                            await message.channel.send("취소 되었습니다.")
                            break
                        
                    except asyncio.exceptions.TimeoutError:
                        await msg.delete()
                        await message.channel.send("시간이 초과되었습니다")

            file.save("memberlist.xlsx")

        elif message.content == "!탈퇴":
            file = openpyxl.load_workbook("memberlist.xlsx")
            wb = file.active
    
            for i in range(1, 101):
                if wb["A" + str(i)].value == str(message.author.id):
                    embed = discord.Embed(title="탈퇴를 진행합니다", description="아래의 이모지에 반응하세요.",color=0xff9900)
                    msg = await message.channel.send(embed=embed)

                    await msg.add_reaction(yes)
                    await msg.add_reaction(no)

                    try:
                        reaction, user = await client.wait_for("reaction_add", check=check, timeout=15)
                
                        if str(reaction.emoji) == yes:
                            await msg.delete()
                            wb.delete_rows(i)
                            await message.channel.send("탈퇴처리가 정상적으로 완료되었습니다.")
                            break

                        elif str(reaction.emoji) == no:
                            await msg.delete()
                            await message.channel.send("취소 되었습니다.")
                            break

                    except asyncio.exceptions.TimeoutError:
                        await msg.delete()
                        await message.channel.send("시간이 초과되었습니다")
                        break

                elif wb["A" + str(i)].value == None:
                    await message.channel.send("가입하지 않은 사용자입니다.")
                    break

            file.save("memberlist.xlsx")

        elif message.content=="!내정보":
            file = openpyxl.load_workbook("memberlist.xlsx")
            wb = file.active
            file1=openpyxl.load_workbook("quiz.xlsx")
            wb1=file1.active
            for i in range(1, 101):
                if wb["A" + str(i)].value == str(message.author.id):
                    emve=discord.Embed(title="내정보",color=0xff9900)
                    emve.add_field(name="이름",value=message.author.name, inline=True)
                    emve.add_field(name="서버닉네임",value=message.author.display_name,inline=False)
                    emve.add_field(name="가입일",value=wb["C" + str(i)].value)
                    if wb1["A" + str(1)].value==str(message.author):
                        emve.add_filed(name="이번퀴즈의 정답자!!!",value="다음 퀴즈도 열심히 해주세요",intline=False)

                    await message.channel.send(embed=emve)
                    e=1
                    break

            if e==0:
                await message.channel.send("가입하지 않은 사용자입니다.")
            else:
                e=0

        elif message.content.startswith("!청소"):
            try:
                amount = message.content[4:]
                await message.channel.purge(limit=int(amount))
                await message.channel.send(f"**{amount}**개의 메시지를 지웠습니다.")
            except ValueError:
                await message.channel.send("청소하실 메시지의 **수**를 입력해 주세요.")

        elif message.content.startswith("!날씨"):
            
            url="https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query=%EC%A0%95%EC%9E%903%EB%8F%99+%EB%82%A0%EC%94%A8&oquery=%EC%88%98%EC%9B%90+%EB%82%A0%EC%94%A8&tqi=hfXOjlprvh8ssflQbbVssssstI8-011814"
            res=requests.get(url)
            res.raise_for_status()
            soup=BeautifulSoup(res.text,"lxml")
            cast=soup.find("p",attrs={"class":"cast_txt"}).get_text()

            curr_temp = soup.find("p",attrs={"class":"info_temperature"}).get_text().replace("도씨","")
            min_temp = soup.find("span",attrs={"class":"min"}).get_text()
            max_temp = soup.find("span",attrs={"class":"max"}).get_text()

            morning_rain_rate=soup.find("span",attrs={"class":"point_time morning"}).get_text().strip()
            afternoon_rain_rate=soup.find("span",attrs={"class":"point_time afternoon"}).get_text().strip()
        
            emved=discord.Embed(color=0xff9900)
            emved.add_field(name="[오늘의 날씨]",value=cast,inline=False)
            emved.add_field(name="현재", value=curr_temp+"",inline=False)
            emved.add_field(name="최저기온", value=min_temp+"C",inline=False)
            emved.add_field(name="최고기온", value=max_temp+"C",inline=False)
            emved.add_field(name="오전강수확률", value=morning_rain_rate,inline=True)
            emved.add_field(name="오후강수확률", value=afternoon_rain_rate,inline=False)
            
        
            await message.channel.send(embed=emved)
        
        elif message.content=="!돈받기":
            file = openpyxl.load_workbook("memberlist.xlsx")
            wb = file.active
            for i in range(1, 101):
                if wb["A" + str(i)].value == str(message.author.id):
                    money=money+500
                    wb["D" + str(i)].value = money
                   
                    e=0 
                    break                  
                else:
                    e=1

            if e==0:
                await message.channel.send("돈 500을 지급했습니다.")  
                await message.channel.send("현재 돈: "+ str(money)) 
            else:
                await message.channel.send("가입하지 않은 사용자입니다.")

            file.save("memberlist.xlsx")
        
        elif message.content=="!돈":
            file = openpyxl.load_workbook("memberlist.xlsx")
            wb = file.active
            for i in range(1, 101):
                if wb["A" + str(i)].value == str(message.author.id):
                    wb["D" + str(i)].value = money
                    e=0
                    break
                else:
                    e=1
            if e==0:
                await message.channel.send(wb["D" + str(i)].value) 
            else:
                await message.channel.send("가입하지 않은 사용자입니다.") 
            e=0

        


client.run(access_token)

